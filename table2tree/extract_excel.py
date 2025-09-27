import re
import os
import glob
import json
import time
import openpyxl
import requests

import pandas as pd
import numpy as np
import tqdm as tqdm
from loguru import logger

from utils.api_utils import *
from utils.sheet_utils import *
from utils.split_utils import *
from utils.attr_extraction import *
from utils.constants import *


def match_minimal_table_structure(sheet):
    """
    匹配最小结构，即一个或两个单元格的时候
    
    返回 dict: {"content1": "content2"}
    """
    nrows = sheet.max_row
    ncols = sheet.max_column

    first_cell = sheet.cell(row=1, column=1)
    _, _, x2, y2 = get_merge_cell_size(sheet, first_cell.coordinate)

    if x2 == nrows and y2 == ncols: # 一个整体的单元格
        return {get_merge_cell_value(sheet, first_cell.coordinate): None}
    if x2 == nrows: # 横着两个单元格
        second_cell = sheet.cell(row=1, column=y2 + 1)
        _, _, xx2, yy2 = get_merge_cell_size(sheet, second_cell.coordinate)
        if yy2 == ncols:
            return {
                get_merge_cell_value(
                    sheet, first_cell.coordinate
                ): get_merge_cell_value(sheet, second_cell.coordinate)
            }
    if y2 == ncols: # 竖着两个单元格
        second_cell = sheet.cell(row=x2 + 1, column=1)
        _, _, xx2, yy2 = get_merge_cell_size(sheet, second_cell.coordinate)
        if xx2 == nrows:
            return {
                get_merge_cell_value(
                    sheet, first_cell.coordinate
                ): get_merge_cell_value(sheet, second_cell.coordinate)
            }


def match_attr(sheet, pos_list):
    return_dict = {}

    nrows = sheet.max_row
    ncols = sheet.max_column

    x = 1
    while x <= nrows:
        y = 1
        while y <= ncols:
            cell = sheet.cell(row=x, column=y)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            if in_pos_list(x, y, pos_list) is not None:  # 是Schema
                key = get_merge_cell_value(sheet, cell.coordinate)
                if in_pos_list(x, y2 + 1, pos_list) is None:  # 若右边不是Schema
                    return_dict[key] = get_merge_cell_value(
                        sheet, sheet.cell(row=x, column=y2 + 1).coordinate
                    )
                if in_pos_list(x2 + 1, y, pos_list) is None:  # 若下面不是Schema
                    if key not in return_dict or (
                        key in return_dict
                        and (return_dict[key] is None or return_dict[key] == "")
                    ):
                        return_dict[key] = get_merge_cell_value(
                            sheet, sheet.cell(row=x2 + 1, column=y).coordinate
                        )
            y += y2 - y1 + 1
        x += 1

    return return_dict


def match_list_column(schema, sheet):
    """属性名在表格左侧的情况下提取子半结构化表格"""
    # Step1 获取Schema
    # 找到第一行中高度最高的单元格，即确定整个Schema行合并了多少行
    nrows = sheet.max_row
    ncols = sheet.max_column

    # Step2 提取Schema
    schema = extract_schema_column(schema)
    flattenned_schema = flatten_schema(schema)  # 展平Schema

    # Step3 预处理切分表格，保证从左向右粒度不增，并记录切分的实体关系
    content = sheet
    col = 1
    while col <= ncols:
        x1, y1, x2, y2 = get_merge_cell_size(
            content, content.cell(row=1, column=col).coordinate
        )
        max_width = y2 - y1 + 1
        row = x2 + 1
        while row <= nrows:
            cell = content.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(content, cell.coordinate)
            width = y2 - y1 + 1
            if width > max_width:  # split
                value = content.cell(row=x1, column=y1).value
                content.unmerge_cells(get_coordinate_by_cell_pos(x1, y1, x2, y2))
                content[
                    get_coordinate_by_cell_pos(x1, y1 + max_width, x1, y1 + max_width)
                ] = value
                content.merge_cells(
                    get_coordinate_by_cell_pos(x1, y1, x2, y1 + max_width - 1)
                )
                content.merge_cells(
                    get_coordinate_by_cell_pos(x1, y1 + max_width, x2, y2)
                )
            row += x2 - x1 + 1
        col += max_width

    # Step4 开始递归的按列切分表格
    col_content = extract_columns(content)

    print(flattenned_schema, col_content)

    # Step5 构建嵌套的字典格式
    return_list, _ = schema_content_match(flattenned_schema, col_content)

    return return_list


def get_schema_height(sheet, direction):
    height = 1

    nrows = sheet.max_row
    ncols = sheet.max_column

    if direction == SCHEMA_TOP:
        for col in range(1, ncols + 1):
            _, _, x2, _ = get_merge_cell_size(
                sheet, sheet.cell(row=1, column=col).coordinate
            )
            height = max(height, x2)

    elif direction == SCHEMA_LEFT:
        for row in range(1, nrows + 1):
            _, _, _, y2 = get_merge_cell_size(
                sheet, sheet.cell(row=row, column=1).coordinate
            )
            height = max(height, y2)

    return height

def get_nrow_cells(sheet):
    n = 0
    
    nrows = sheet.max_row
    ncols = sheet.max_column
    
    for row in range(1, nrows + 1):
        col = 1
        cnt = 0
        while col <= ncols:
            cnt += 1            
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            col += (y2 - y1 + 1)
        n = max(n, cnt)
    return n

def get_ncol_cells(sheet):
    n = 0
    
    nrows = sheet.max_row
    ncols = sheet.max_column
    
    for col in range(1, ncols + 1):
        row = 1
        cnt = 0
        while row <= nrows:
            cnt += 1
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            row += (x2 - x1 + 1)
        n = max(n, cnt)
    
    return n

def preprocess_cell(value):
    value = str(value)
    # value = re.sub(r"\s+", "", value)  # 去除所有空白字符
    return value


def preprocess_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.value = preprocess_cell(cell.value)
    return sheet


def get_xlsx_sheet(file):
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = preprocess_sheet(wb.active)
    return sheet


def get_structured_xlsx_sheet(file):
    wb = openpyxl.load_workbook(file, data_only=True)

    sheet = wb.active

    sheet = sheet2structure(sheet)

    sheet = preprocess_sheet(sheet)
    
    return sheet


################################# Main #################################
def process_sheet_vlm(sheet,            # 待处理的Excel工作簿格式的表格
                      get_json=False,   # 获得VLM + Rule识别后的中间结果JSON
                      cache=False       # 保存中间结果文件
                      ):
    
    ##### Step 0: 确认 sheet 非空
    if sheet is None:
        logger.info(f'Sheet 为空, process_sheet_vlm() 直接返回')

    ##### Step 1: 删除空行，空列 Tested
    delete_empty_columns(sheet)
    delete_empty_rows(sheet)

    nrows = sheet.max_row
    ncols = sheet.max_column

    nrow_cells = get_nrow_cells(sheet)
    ncol_cells = get_ncol_cells(sheet)

    logger.info(f'{DELIMITER} 进入 process_sheet_vlm() 函数 {DELIMITER}')
    logger.info(f'表格行数: {nrows} 表格列数: {ncols}\n表格每行最多出现的Cell数: {nrow_cells} 表格每列最多出现的Cell数: {ncol_cells}')

    ##### 递归返回条件 1: 匹配到最小结构单元
    res = match_minimal_table_structure(sheet)
    logger.info(f"{DELIMITER} 递归返回条件1: 匹配到最小结构单元 {DELIMITER}")
    if res is not None:
        logger.info(f"匹配到最小结构单元: {res}")
        return res
    else:
        logger.info(f"未匹配到最小结构单元")

    ##### 递归返回条件 2: 表为小表
    logger.info(f"{DELIMITER} 递归返回条件2: 表为小表 {DELIMITER}")
    if (nrow_cells < SMALL_TABLE_ROWS and ncol_cells < BIG_TABLE_COLUMNS) or (
        nrow_cells < BIG_TABLE_ROWS and ncol_cells < SMALL_TABLE_COLUMNS
    ):
        res = vlm_get_json(sheet, enable_crop=False)
        logger.info(f"VLM 小表直接解析结果: {res}")
        return res
    else:
        logger.info(f"该表不是小表")

    ##### Step 2: 拆分整行 / 整列的合并单元格
    logger.info(f"{DELIMITER} 拆分整行 / 整列的合并单元格 {DELIMITER}")
    sheet_dict = rowspan_entire(sheet)  # 获得深度可能不为1的sheet_dict
    if len(sheet_dict) == 1 and list(sheet_dict.keys())[0] == DEFAULT_TABLE_NAME:
        sheet_dict = colspan_entire(sheet)
    logger.info(f"拆分出 {len(sheet_dict)} 个子表")
    logger.info(f"拆分结果列表: {sheet_dict}")

    ##### Step 3: 遍历每一个子表进行处理
    for key, sheet in sheet_dict.items():
        if sheet is None: continue

        delete_empty_columns(sheet)
        delete_empty_rows(sheet)

        ##### Step 4: 提取并判断 Schema 的方向

        # 根据表格合并单元格的粒度变化来判断Schema方向
        # if granularity_decrease_col(sheet):
        #     granularity_directon_col = SCHEMA_TOP
        # else:
        #     granularity_directon_col = SCHEMA_FAIL
        # if granularity_decrease_row(sheet):
        #     granularity_directon_row = SCHEMA_LEFT
        # else:
        #     granularity_directon_row = SCHEMA_FAIL

        ##### 使用 VLM 识别 Schema
        logger.info(f"{DELIMITER} 使用 VLM 识别表格 Schema {DELIMITER}")
        schema_vague = vlm_get_schema(sheet)
        pos_list, schema_list, _ = schema_pos_match(
            sheet, schema_vague, enable_embedding=True
        )  # 将 vlm 的输出与表格内容对应，并获取 schema 的位置
        direction = get_schema_direction_by_pos(sheet, pos_list)  # Rule 2
        # direction = get_schema_direction_by_vlm(sheet)
        schema_height = get_schema_height(sheet, direction)
        d = "上部" if direction == SCHEMA_TOP else "左侧"

        logger.info(f"子表格 {key} Schema 的方向为 {d}")
        logger.info(f"子表格 {key} Schema 的高度为 {schema_height}")
        logger.info(f"子表格 {key} Schema 为 {schema_list}")

        ##### Step 5: 根据 schema 方向拆分表格, Schema在左边, 上下拆分为并列的多个子部分; Schema在上面, 左右拆分为并列的多个子部分
        logger.info(f"{DELIMITER} 根据 Schema 方向拆解并列表格 {DELIMITER}")
        parallel_sheet = None

        if direction == SCHEMA_TOP:
            parallel_sheet = split_subtable_row(sheet)
            new_parallel_sheet = []
            for x in parallel_sheet:
                if x is not None:
                    new_parallel_sheet.append(x)
            parallel_sheet = new_parallel_sheet
            logger.info(f"子表 {key} 根据粒度变化拆解为了 {len(parallel_sheet)} 个子部分")

        elif direction == SCHEMA_LEFT:
            # 根据粒度变化判断是一个结构化表格，还是一个半结构化子表
            schema, data_sheet = split_schema_column(sheet)
            if (
                data_sheet is not None and granularity_decrease_row(data_sheet) and data_sheet.max_column / data_sheet.max_row > 2
            ):  # 行粒度递减并且是宽表
                parallel_sheet = []
                sheet = transpose_sheet(sheet)
            else:
                parallel_sheet = split_subtable_each_row(sheet, schema_height)
                logger.info(f"子表 {key} 根据行数拆解为了 {len(parallel_sheet)} 个子部分")
        
        else:
            res = vlm_get_json(sheet, save=cache, enable_crop=False)
            sheet_dict[key] = res
            
            logger.info(f"子表 {key} 识别 Schema 方向失败, 使用 VLM 直接解析为 JSON")
            logger.info(f"{res}")
            continue

        # Step 6 如果sheet被分为了多个并列的sheet，则分别递归处理，否则直接处理
        logger.info(f"{DELIMITER} 递归处理平行表格 {DELIMITER}")
        if len(parallel_sheet) > 1: # 分别递归处理，否则直接处理
            logger.info(f"递归处理多张平行表格")
            if isinstance(parallel_sheet, list):
                json_dict = {}
                subtable_cnt = 1                
                for index, subsheet in enumerate(parallel_sheet):
                    if subsheet is None:
                        logger.info(f"平行表格 {subtable_cnt} 默认表名 {DEFAULT_SUBTABLE_NAME}{subtable_cnt} 为None")
                        continue
                    logger.info(f"开始递归处理平行表格 {subtable_cnt} 默认表名 {DEFAULT_SUBTABLE_NAME}{subtable_cnt}")
                    
                    res = process_sheet_vlm(subsheet, get_json)
                    if DEFAULT_TABLE_NAME in res:
                        res = res[DEFAULT_TABLE_NAME]
                        if isinstance(res, dict):
                            json_dict.update(res)
                        else:
                            json_dict[f"{DEFAULT_SUBTABLE_NAME}{subtable_cnt}"] = res
                            subtable_cnt += 1
                    else:
                        json_dict.update(res)
                    logger.info(f"完成递归处理平行表格 {subtable_cnt} 默认表名 {DEFAULT_SUBTABLE_NAME}{subtable_cnt}")
                    
                sheet_dict[key] = json_dict
            elif isinstance(parallel_sheet, dict):
                json_dict = {}
                for name, sheet in parallel_sheet.items():
                    logger.info(f"开始递归处理平行表格 指定表名 {name}")
                    if isinstance(sheet, (str, int, float)) or sheet is None:
                        logger.info(f"isinstance(sheet, (str, int, float)) or sheet is None")
                        json_dict[name] = sheet
                    else:
                        res = process_sheet_vlm(sheet, get_json)  # a dict
                        if len(res) == 1 and DEFAULT_TABLE_NAME in res:
                            res = res[DEFAULT_TABLE_NAME]
                        if DEFAULT_SUBTABLE_NAME in name and isinstance(res, dict):
                            json_dict.update(res)
                        else:
                            json_dict[name] = res
                    logger.info(f"完成递归处理平行表格 指定表名 {name}")
                sheet_dict[key] = json_dict

        else:   # 直接处理
            logger.info(f"只有一张平行表格，直接进行处理")
            
            if get_json:
                try:
                    schema_sheet, data_sheet = split_schema_row(sheet)
                    res = match_list_column(schema_sheet, data_sheet)
                except Exception as e:
                    import traceback; traceback.print_exc()
                    res = vlm_get_json(sheet, save=cache, enable_crop=False)
                sheet_dict[key] = res

    return sheet_dict


def process_table_vlm(file,              # 输入的表格文件路径
                     get_json=False,    # 是否获取 JSON
                     cache=False        # 是否保存中间结果
                     ):

    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = preprocess_sheet(wb.active)
    res = process_sheet_vlm(sheet, get_json=get_json, cache=cache)

    return res
