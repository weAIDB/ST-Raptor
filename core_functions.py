import os
import json
import time
import shutil
import pickle
import types
import openpyxl
import pandas as pd
import re
from datetime import datetime
import gradio as gr
from loguru import logger
from utils.constants import DELIMITER, LOG_DIR
from embedding import EmbeddingModel
from utils.api_utils import vlm_generate, llm_generate, embedding_generate
from query.primitive_pipeline import *
from table2tree.feature_tree import *
from table2tree.extract_excel import process_sheet_vlm, preprocess_sheet
from config import api_config
from utils.sheet_utils import html2workbook, extract_markdown_tables

# 全局思维链条数据存储
thinking_chain_data = {}

def save_tree_artifacts(f_tree, cache_dir):
    """Save all tree artifacts to the specified directory"""
    tree_json = f_tree.__json__()
    tree_str = f_tree.__str__([1])
    
    with open(os.path.join(cache_dir, "temp.pkl"), "wb") as f:
        pickle.dump(f_tree, f)
    with open(os.path.join(cache_dir, "temp.txt"), "w", encoding='utf-8') as f:
        f.write(tree_str)
    with open(os.path.join(cache_dir, "temp.json"), "w", encoding='utf-8') as f:
        json.dump(tree_json, f, indent=4, ensure_ascii=False)
    
    # Generate and save embeddings
    try:
        raw_values = f_tree.all_value_list()
        texts = [str(x) for x in raw_values] if raw_values else []
        if texts:
            embedding_dict = EmbeddingModel().get_embedding_dict(texts)
            EmbeddingModel().save_embedding_dict(
                embedding_dict, os.path.join(cache_dir, "temp.embedding.json")
            )
    except Exception as ee:
        logger.error(f"embedding generate failed: {ee}")

def ensure_cache_directories(cache_dir, temp_dir=None):
    """Ensure cache directories exist"""
    os.makedirs(cache_dir, exist_ok=True)
    if temp_dir:
        os.makedirs(temp_dir, exist_ok=True)

def handle_processing_error(e, error_prefix="处理"):
    """Standard error handling for processing functions"""
    import traceback
    error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
    gr.Warning(f"❌ {error_prefix}失败: {error_msg}")
    return f"{error_prefix}失败"

def setup_cache_directory(conversation_id, default_cache_dir="cache", default_temp_dir="data/SSTQA/temp_tables"):
    """Setup cache directory based on conversation_id"""
    if conversation_id:
        cache_dir = os.path.join("history", conversation_id)
        temp_dir = os.path.join("history", conversation_id)
    else:
        cache_dir = default_cache_dir
        temp_dir = default_temp_dir
    
    os.makedirs(cache_dir, exist_ok=True)
    os.makedirs(temp_dir, exist_ok=True)
    
    return cache_dir, temp_dir


def save_placeholder_data(cache_dir, placeholder_data, embedding_texts=None):
    """Save placeholder data and embeddings for conversation history"""
    try:
        # Save placeholder pkl file
        with open(os.path.join(cache_dir, "temp.pkl"), "wb") as f:
            pickle.dump(placeholder_data, f)
        
        # Generate and save embeddings if provided
        if embedding_texts:
            embedding_dict = EmbeddingModel().get_embedding_dict(embedding_texts)
            EmbeddingModel().save_embedding_dict(
                embedding_dict, os.path.join(cache_dir, "temp.embedding.json")
            )
    except Exception as e:
        logger.error(f"Failed to save placeholder data: {e}")


def ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts=None):
    """Ensure conversation cache directory exists and save placeholder data"""
    if conversation_id:
        cache_dir = os.path.join("history", conversation_id)
        os.makedirs(cache_dir, exist_ok=True)
        save_placeholder_data(cache_dir, placeholder_data, embedding_texts)
        return cache_dir
    return None


def clear_directory_contents(dir_path):
    """Clear all contents from a directory, preserving the directory itself"""
    if os.path.exists(dir_path):
        for item in os.listdir(dir_path):
            item_path = os.path.join(dir_path, item)
            if os.path.isfile(item_path):
                os.remove(item_path)
            elif os.path.isdir(item_path):
                shutil.rmtree(item_path)  # 递归删除子目录

def generate_conversation_id():
    """生成唯一的对话ID，基于时间戳"""
    return datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # 包含毫秒以确保唯一性

def _contains_cjk(text):
    return bool(re.search(r"[\u4e00-\u9fff]", text or ""))


def _truncate_title(title, max_words=8, max_chars=40):
    title = (title or "").strip()
    if not title:
        return ""
    words = title.split()
    if len(words) > max_words:
        title = " ".join(words[:max_words])
    if len(title) > max_chars:
        title = title[:max_chars].rstrip() + "..."
    return title


def generate_history_title_from_questions(chat_history):
    """Generate a concise English history title from user questions."""
    if not chat_history or not isinstance(chat_history, list):
        return None
    
    # 提取所有用户问题
    user_questions = []
    for msg in chat_history:
        if isinstance(msg, dict) and msg.get("role") == "user":
            question = msg.get("content", "").strip()
            if question:
                user_questions.append(question)
    
    if not user_questions:
        return None
    
    # 如果只有一个问题，直接使用它；如果有多个，用LLM概括
    if len(user_questions) == 1:
        question_text = user_questions[0]
    else:
        # 合并所有问题
        questions_text = "\n".join([f"{i+1}. {q}" for i, q in enumerate(user_questions)])
        question_text = questions_text
    
    # 限制问题文本长度，避免超出LLM上下文
    if len(question_text) > 500:
        question_text = question_text[:500] + "..."
    
    # 使用LLM生成标题
    prompt = f"""Generate a concise English title (no more than 8 words) that summarizes the core topic of the conversation.

User questions:
{question_text}

Return the title only. Do not add explanations or quotes."""
    
    try:
        title = get_llm_generate(prompt, max_tokens=50, temperature=0.3)
        title = title.strip().strip('"').strip("'")
        title = _truncate_title(title, max_words=8, max_chars=40)
        if _contains_cjk(title):
            return "Conversation Summary"
        return title if title else None
    except Exception as e:
        logger.error(f"生成历史记录标题失败: {e}")
        return None

def create_conversation_record(conversation_id, file_list, upload_time, summary, chat_history=None):
    """创建对话记录文件，用于历史记录显示
    
    参数:
    conversation_id: 对话ID
    file_list: 文件列表
    upload_time: 上传时间
    summary: 摘要（如果提供chat_history且有用户问题，会用LLM生成标题覆盖summary）
    chat_history: 对话历史，格式为[{"role": "user", "content": "问题"}, ...]
    """
    history_dir = "history"
    os.makedirs(history_dir, exist_ok=True)
    
    record_file = os.path.join(history_dir, "history_records.json")
    
    # 读取现有记录
    records = []
    if os.path.exists(record_file):
        try:
            with open(record_file, 'r', encoding='utf-8') as f:
                records = json.load(f)
        except:
            records = []
    
    # 如果有对话历史且有用户问题，尝试用LLM生成标题
    final_summary = summary
    if chat_history:
        llm_title = generate_history_title_from_questions(chat_history)
        if llm_title:
            final_summary = llm_title
        # 如果没有生成标题，使用默认summary
    final_summary = _truncate_title(final_summary, max_words=8, max_chars=40) or "Conversation Summary"
    
    # 添加新记录
    new_record = {
        "conversation_id": conversation_id,
        "file_list": file_list,
        "upload_time": upload_time,
        "summary": final_summary
    }
    
    records.append(new_record)
    
    # 保存记录
    with open(record_file, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def get_conversation_records():
    """获取所有对话记录，用于历史记录显示"""
    history_dir = "history"
    os.makedirs(history_dir, exist_ok=True)
    
    record_file = os.path.join(history_dir, "history_records.json")
    
    # 读取记录
    records = []
    if os.path.exists(record_file):
        try:
            with open(record_file, 'r', encoding='utf-8') as f:
                records = json.load(f)
        except:
            records = []
    
    # 转换为表格格式数据
    table_data = []
    for record in records:
        # 将文件列表转换为字符串
        file_names_str = ", ".join(record.get("file_list", []))
        table_data.append([
            record.get("conversation_id", ""),
            file_names_str,
            record.get("upload_time", ""),
            record.get("summary", ""),
            "查看"  # 操作列
        ])
    
    return table_data

def get_llm_generate(prompt, max_tokens=8192, temperature=0.5):
    return llm_generate(
        prompt=prompt,
        key=api_config["llm_api_key"],
        url=api_config["llm_api_url"],
        model=api_config["llm_model"],
        max_tokens=max_tokens,
        temperature=temperature
    )

def reshape_question_with_context(current_question, chat_history, temperature=0.5):
    """
    使用上下文重塑用户问题，明确并代替可能指代不明确的代词
    
    参数:
    current_question: 当前用户问题
    chat_history: 对话历史，格式为[{"role": "user", "content": "问题"}, {"role": "assistant", "content": "回答"}, ...]
    temperature: LLM温度参数
    
    返回:
    重塑后的清晰问题
    """
    if not chat_history or len(chat_history) == 0:
        return current_question
    
    # 构建上下文提示
    context_prompt = "对话历史:\n"
    for message in chat_history:
        role = "用户" if message["role"] == "user" else "助手"
        context_prompt += f"{role}: {message['content']}\n"
    
    context_prompt += f"\n当前问题: {current_question}\n"
    context_prompt += "\n请根据对话历史，重塑当前问题，明确并代替可能指代不明确的代词，保持问题的核心意思不变。"
    context_prompt += "\n重塑后的问题:"
    
    try:
        reshaped_question = get_llm_generate(
            prompt=context_prompt,
            max_tokens=256,
            temperature=temperature
        )
        logger.info(f"原始问题: {current_question}")
        logger.info(f"重塑问题: {reshaped_question}")
        return reshaped_question.strip()
    except Exception as e:
        logger.error(f"问题重塑失败: {str(e)}")
        # 如果重塑失败，返回原始问题
        return current_question

def get_vlm_generate():
    # 返回一个已经配置好API参数的vlm_generate函数
    def configured_vlm_generate(prompt, image, temperature=0.5):
        return vlm_generate(
            prompt=prompt,
            image=image,
            key=api_config["vlm_api_key"],
            url=api_config["vlm_api_url"],
            model=api_config["vlm_model"],
            temperature=temperature
        )
    return configured_vlm_generate

def get_embedding_generate():
    # 返回一个已经配置好API参数的embedding_generate函数
    def configured_embedding_generate(input_texts, dimensions=1024):
        return embedding_generate(
            input_texts=input_texts,
            key=api_config["embedding_api_key"],
            url=api_config["embedding_api_url"],
            model=api_config["embedding_model"],
            dimensions=dimensions
        )
    return configured_embedding_generate


def convert_to_xlsx(src_path, dest_path):
    """将各种格式的文件转换为 xlsx 格式"""
    ext = os.path.splitext(src_path)[1].lower()
    try:
        if ext == ".xlsx":
            shutil.copy2(src_path, dest_path)
        elif ext == ".csv":
            df_src = pd.read_csv(src_path)
            df_src.to_excel(dest_path, index=False, engine="openpyxl")
        elif ext == ".html":
            html_content = open(src_path, "r", encoding="utf-8").read()
            html2workbook(html_content).save(dest_path)
        elif ext == ".md":
            md_content = open(src_path, "r", encoding="utf-8").read()
            table = extract_markdown_tables(md_content)
            if table and len(table) > 1:
                df_src = pd.DataFrame(table[1:], columns=table[0])
                df_src.to_excel(dest_path, index=False, engine="openpyxl")
            else:
                shutil.copy2(src_path, dest_path)
        else:
            shutil.copy2(src_path, dest_path)
    except Exception as e:
        logger.error(f"转换文件 {src_path} 到 xlsx 失败: {e}")
        shutil.copy2(src_path, dest_path)


def get_multiple_excel_feature_tree(files, log_dir=LOG_DIR, vlm_cache=False):
    """处理多个 Excel 文件，并构建成一棵总树，根节点为 'alldocument'"""
    all_docs_dict = {}
    temp_dir = "data/SSTQA/temp_tables"
    os.makedirs(temp_dir, exist_ok=True)
    
    for i, file_obj in enumerate(files):
        # file_obj 可能是 Gradio 的 File 对象或 SimpleNamespace
        src_path = file_obj.name if hasattr(file_obj, 'name') else str(file_obj)
        filename = os.path.basename(src_path)
        
        # 为每个文件创建一个唯一的临时 xlsx 名
        temp_file = os.path.join(temp_dir, f"temp_{i}.xlsx")
        try:
            convert_to_xlsx(src_path, temp_file)
            
            # 开启处理逻辑
            wb = openpyxl.load_workbook(temp_file, data_only=True)
            file_tree_dict = {}
            for sheet_name in wb.sheetnames:
                logger.info(f"正在处理文件 {filename} 的 Sheet: {sheet_name}")
                sheet = preprocess_sheet(wb[sheet_name])
                # 获取该 sheet 的结构字典 (tree_dict)
                sheet_tree_dict = process_sheet_vlm(sheet, get_json=False, cache=vlm_cache)
                file_tree_dict[sheet_name] = sheet_tree_dict
            
            # 将该文件的所有 sheet 挂在文件名节点下
            all_docs_dict[filename] = file_tree_dict
        except Exception as e:
            logger.error(f"处理文件 {filename} 失败: {e}")
            continue

    # 构建带 'alldocument' 根节点的字典
    combined_tree_dict = {"alldocument": all_docs_dict}
    
    # 建树并打标签
    total_tree = construct_feature_tree(combined_tree_dict)
    total_tree = tag_feature_tree(total_tree)
    
    return total_tree


def process_multiple_tables_for_tree(files, conversation_id=None):
    """专门处理多个表格，生成统一的 H-OTree 结构"""
    global thinking_chain_data
    thinking_chain_data = {"question_answering": {}, "retrieval_chains": []}
    
    if not files:
        return None
    
    try:
        # 如果提供了 conversation_id，则创建专用文件夹，否则使用 cache 目录
        if conversation_id:
            cache_dir = os.path.join("history", conversation_id)
            os.makedirs(cache_dir, exist_ok=True)
        else:
            cache_dir = "cache"
            os.makedirs(cache_dir, exist_ok=True)
        
        log_dir = LOG_DIR
        
        # 处理表格生成总树
        start_time = time.time()
        f_tree = get_multiple_excel_feature_tree(files, log_dir=log_dir, vlm_cache=False)
        tree_json = f_tree.__json__()
        tree_str = f_tree.__str__([1])
        end_time = time.time()
        
        # 保存中间文件 (使用与单文件一致的名称，供问答逻辑使用)
        save_tree_artifacts(f_tree, cache_dir)
            
        return tree_json
    except Exception as e:
        import traceback
        logger.error(f"多文件处理失败: {traceback.format_exc()}")
        return None


def analyze_multiple_files_for_route(files):
    """分析多个文件以确定处理线路"""
    if not files:
        return "请选择文件"
    
    # 分析所有文件
    has_image = False
    has_xlsx = False
    has_text = False
    file_details = []
    
    for file in files:
        file_path = file.name if hasattr(file, 'name') else file
        ext = os.path.splitext(file_path)[1].lower()
        
        if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
            has_image = True
        elif ext in [".xlsx", ".xls", ".docx", ".doc"]:
            has_xlsx = True
        else:
            has_text = True  # 包括 .txt, .md, .json, .csv 等
        
        file_size = os.path.getsize(file_path)
        file_details.append({
            "path": file_path,
            "size": file_size,
            "ext": ext
        })
    
    # 按照优先级判断处理线路
    # 1. 如果有任何图片文件，所有文件一起走VLM
    if has_image:
        return "vlm"
    # 2. 如果有xlsx文件和纯文本内容，走HOTree
    elif has_xlsx:
        return "hotree"
    # 3. 如果只有纯文本文件，走LLM
    elif has_text:
        return "llm"
    else:
        return "llm"  # 默认


def determine_processing_route(file_path, file_size, file_content=None):
    """使用AI判断文件处理线路"""
    if not file_path:
        return "请选择文件"
    
    # 获取文件扩展名
    ext = os.path.splitext(file_path)[1].lower()
    
    # 构建提示词
    prompt = f"文件路径: {file_path}\n"
    prompt += f"文件大小: {file_size} 字节\n"
    prompt += f"文件类型: {ext}\n"
    if file_content:
        prompt += f"文件内容摘要: {file_content[:500]}...\n"
    prompt += "请根据以上信息判断应该使用哪种处理线路：\n"
    prompt += "1. 'llm'：纯文本内容，适合使用LLM处理\n"
    prompt += "2. 'vlm'：包含图片或需要视觉理解的内容，适合使用VLM处理\n"
    prompt += "3. 'hotree'：结构化数据或表格内容，适合使用H-OTree处理\n"
    prompt += "请只返回'llm'、'vlm'或'hotree'中的一个，不要添加任何其他解释。"
    
    try:
        # 调用LLM生成判断结果
        result = get_llm_generate(prompt, max_tokens=10, temperature=0.1)
        result = result.strip().lower()
        
        # 验证结果有效性
        if result in ["llm", "vlm", "hotree"]:
            return result
        else:
            # 如果AI返回无效结果，使用默认规则
            if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
                return "vlm"
            elif ext in [".xlsx", ".xls", ".docx", ".doc"]:
                return "hotree"
            else:
                return "llm"
    except Exception as e:
        # 如果AI调用失败，使用默认规则
        logger.error(f"AI判断线路失败: {e}")
        if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
            return "vlm"
        elif ext in [".xlsx", ".xls", ".docx", ".doc"]:
            return "hotree"
        else:
            return "llm"

def answer_question(
    qa_pair: dict,                          # 一条问答对
    table_file: str,                        # 表格原文件路径
    cache_dir: str,                           # 存储 HO-Tree 中间结果的路径
    enable_query_decompose: bool = True,    # 是否启用 Query Decomposition 机制
    enable_emebdding: bool = True,          # 是否启用 Embedding 机制
    log_dir: str = LOG_DIR,                 # Log 日志目录
    temperature: float = 0.5,               # LLM/VLM temperature
    max_tokens: int = 2048                  # LLM/VLM max_tokens
):
    
    query = qa_pair["query"]

    ##### 创建日志文件 命名为 表格id_问题id.log
    log_file = os.path.join(log_dir, f'temp.log')
    log_file_handler = logger.add(
        log_file,
        enqueue=False,  # 不使用队列，立即写入，避免缓冲
        backtrace=False,
        diagnose=False
    )

    logger.info(f"{DELIMITER} 开始问答问题 {DELIMITER}")

    start_time = time.time()

    logger.info(f"Question ID: temp")
    logger.info(f"Table ID: temp")

    logger.info(f"Question: {query}")
    logger.info(f"Temperature: {temperature}")
    logger.info(f"Max tokens: {max_tokens}")

    ##### 加载 ho_tree
    pkl_file = os.path.join(cache_dir, f'temp.pkl')
    embedding_cache_file = os.path.join(cache_dir, f'temp.embedding.json')
    with open(pkl_file, 'rb') as file:
        ho_tree = pickle.load(file)

    logger.info(f"Loading PKL File: {pkl_file}")
    logger.info(f"Loading Embedding Cache File: {embedding_cache_file}")

    final_answer, _, reliability = qa_RWP(
        query=query,
        ho_tree=ho_tree,
        table_file=table_file,
        embedding_cache_file=embedding_cache_file,
        enable_emebdding=enable_emebdding,
        enable_query_decompose=enable_query_decompose,
        temperature=temperature,
        max_tokens=max_tokens
    )
    qa_pair["reliability"] = reliability
    qa_pair["model_output"] = final_answer

    end_time = time.time()

    logger.info(f"{DELIMITER} 回答问题成功！ {DELIMITER}")
    logger.info(f"Cost time: {end_time - start_time}")
    
    logger.remove(log_file_handler)
    
    return qa_pair

def get_excel_feature_tree_multisheet(file: str,                   # 输入表格文件路径
                                     log_dir: str = LOG_DIR,      # LOG 日志记录路径
                                     vlm_cache: bool = False      # 是否保存转图片的中间结果
                                     ):
    """处理 Excel 文件中的所有 sheet，并构建成一棵总树"""
    # 1. 打开文件获取所有 sheet
    wb = openpyxl.load_workbook(file, data_only=True)
    combined_tree_dict = {}
    
    # 2. 循环处理每一个 sheet
    for sheet_name in wb.sheetnames:
        logger.info(f"正在处理 Sheet: {sheet_name}")
        sheet = preprocess_sheet(wb[sheet_name])
        # 获取该 sheet 的结构字典 (tree_dict)
        sheet_tree_dict = process_sheet_vlm(sheet, get_json=False, cache=vlm_cache)
        
        # 将每个 sheet 挂在以 sheet_name 命名的节点下
        combined_tree_dict[sheet_name] = sheet_tree_dict

    # 3. 传入大字典，一键生成多层级的总树
    # construct_feature_tree 会递归处理字典
    total_tree = construct_feature_tree(combined_tree_dict)
    
    # 4. 递归打标签
    total_tree = tag_feature_tree(total_tree)
    
    return total_tree


def process_table_for_tree(file, conversation_id=None):
    """专门处理表格，生成H-OTree结构"""
    global thinking_chain_data
    
    # 重置思维链条数据
    thinking_chain_data = {
        "question_answering": {},
        "retrieval_chains": []
    }
    
    if file is None:
        return "请先选择表格文件", ""
    
    # 注意：这里不再自动调用 clear_all()，由外部调用者根据需要决定
    try:
        # 设置缓存目录
        cache_dir, temp_dir = setup_cache_directory(conversation_id, "cache", "data/SSTQA/temp_tables")
        source_filename = os.path.splitext(os.path.basename(file.name))[0]
        
        # 定义日志目录（保持在全局LOG_DIR，用于调试目的）
        log_dir = LOG_DIR
                
        # 创建临时文件
        temp_file = os.path.join(temp_dir, "temp.xlsx")

        # 兼容多种格式，统一转为 xlsx
        src_path = file.name
        ext = os.path.splitext(src_path)[1].lower()
        try:
            if ext == ".xlsx":
                shutil.copy2(src_path, temp_file)
            elif ext == ".csv":
                df_src = pd.read_csv(src_path)
                df_src.to_excel(temp_file, index=False, engine="openpyxl")
            elif ext == ".html":
                html_content = open(src_path, "r", encoding="utf-8").read()
                html2workbook(html_content).save(temp_file)
            elif ext == ".md":
                md_content = open(src_path, "r", encoding="utf-8").read()
                table = extract_markdown_tables(md_content)
                with pd.ExcelWriter(temp_file, engine="openpyxl") as writer:
                    sheet_name = "sheet"
                    df_src = pd.DataFrame(table[1:], columns=table[0])
                    df_src.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                shutil.copy2(src_path, temp_file)
        except Exception as e:
            logger.error(f"格式转换失败: {e}")
            return "文件格式不支持或转换失败", ""
                
        # 读取表格
        df = pd.read_excel(temp_file)
            
        # 处理表格生成H-OTree
        start_time = time.time()
        # 使用多 Sheet 版本处理
        f_tree = get_excel_feature_tree_multisheet(temp_file, log_dir=log_dir, vlm_cache=False)
        tree_json = f_tree.__json__()
        tree_str = f_tree.__str__([1])
        end_time = time.time()
                
        # 保存中间文件
        save_tree_artifacts(f_tree, cache_dir)
        # 保存额外的副本
        with open(os.path.join(cache_dir, f"temp1.json"), "w", encoding='utf-8') as f:
            json.dump(tree_json, f, indent=4, ensure_ascii=False)
        
        # 这里移除 gr.Info，避免在循环处理多文件时产生大量弹窗
        return tree_json
         
    except Exception as e:
        return handle_processing_error(e, "生成树")
    

def pure_llm_generate_answer(question, context="", temperature=0.5, max_tokens=2048):
    """Generate an answer using pure LLM."""
    if not question.strip():
        gr.Warning("Please enter a question")
        return "Please enter a question"
    try:
        # 构建提示词
        prompt = f"Question: {question}\n"
        if context:
            prompt += f"Context: {context}\n"
        prompt += "Please answer the question based on the information above."
        
        # 调用LLM生成答案
        answer = get_llm_generate(prompt, max_tokens, temperature)
        
        gr.Info("✅ LLM answer generated successfully!")
        return f"Answer: {answer}"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ LLM answer generation failed: {error_msg}")
        return "Failed to generate answer"


def pure_vlm_generate_answer(question, image_path, temperature=0.5, max_tokens=2048):
    """Generate an answer using pure VLM."""
    if not question.strip():
        gr.Warning("Please enter a question")
        return "Please enter a question"
    if not image_path:
        gr.Warning("Please select an image file")
        return "Please select an image file"
    try:
        # 构建提示词
        prompt = f"Question: {question}\n"
        prompt += "Please answer the question based on the image."
        
        # 调用VLM生成答案
        vlm_generate_func = get_vlm_generate()
        answer = vlm_generate_func(prompt, image_path, temperature)
        
        gr.Info("✅ VLM answer generated successfully!")
        return f"Answer: {answer}"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ VLM answer generation failed: {error_msg}")
        return "Failed to generate answer"


def process_file_with_route(file, question, temperature=0.5, max_tokens=2048, conversation_id=None):
    """根据文件类型自动选择处理线路，支持单个文件或多个文件"""
    if not file:
        gr.Warning("Please select a file")
        return "Please select a file"
    
    # 检查是否是多个文件
    if isinstance(file, list):
        return process_multiple_files_with_route(file, question, temperature, max_tokens, conversation_id=conversation_id)
    else:
        # 单个文件处理
        try:
            # 获取文件信息
            file_path = file.name
            file_size = os.path.getsize(file_path)
            
            # 读取文件内容摘要
            file_content = None
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    file_content = f.read(1000)
            except:
                # 二进制文件无法读取内容
                pass
            
            # 确定处理线路
            route = determine_processing_route(file_path, file_size, file_content)
            
            # 根据线路处理文件
            if route == "llm":
                # 纯LLM处理
                if not file_content:
                    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                        file_content = f.read()
                
                # 为LLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "text",
                        "file_path": file_path,
                        "file_content_preview": file_content[:500],  # 限制长度
                        "processing_method": "llm"
                    }
                    
                    # 为文本内容生成嵌入向量
                    embedding_texts = [f"Text file: {os.path.basename(file_path)}, content: {file_content[:500]}"]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                return pure_llm_generate_answer(question, file_content, temperature, max_tokens)
            elif route == "vlm":
                # 纯VLM处理
                # 为VLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "image",
                        "file_path": file_path,
                        "processing_method": "vlm"
                    }
                    
                    # 为图像上下文生成简单的嵌入向量
                    embedding_texts = [f"Image file: {os.path.basename(file_path)}"]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                return pure_vlm_generate_answer(question, file_path, temperature, max_tokens)
            elif route == "hotree":
                # H-OTree处理
                wrapped_file = types.SimpleNamespace(name=file.name)
                # process_table_for_tree会将处理结果保存到临时文件，供后续问答使用
                data = process_table_for_tree(wrapped_file, conversation_id=conversation_id)
                if data:
                    # 使用H-OTree方法回答问题
                    result = process_question_only(question, temperature, max_tokens, conversation_id=conversation_id)
                    return result
                else:
                    return "H-OTree处理失败"
            else:
                return f"未知处理线路: {route}"
        except Exception as e:
            import traceback
            error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
            gr.Warning(f"❌ 文件处理失败: {error_msg}")
            return "文件处理失败"


def process_multiple_files_with_route(files, question, temperature=0.5, max_tokens=2048, conversation_id=None):
    """处理多个文件，根据文件类型自动选择处理线路"""
    if not files or len(files) == 0:
        gr.Warning("请选择文件")
        return "请选择文件"
    
    try:
        # 分析多个文件以确定处理线路
        route = analyze_multiple_files_for_route(files)
        
        # 根据线路处理文件
        if route == "llm":
            # 纯LLM处理 - 合并所有文本文件内容
            combined_content = ""
            for file in files:
                file_path = file.name
                try:
                    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                        content = f.read()
                        combined_content += f"\n--- 文件: {os.path.basename(file_path)} ---\n{content}\n"
                except:
                    # 非文本文件跳过或简单描述
                    combined_content += f"\n--- 文件: {os.path.basename(file_path)} (非文本文件) ---\n"
            
            # 为LLM对话创建必要的处理文件，以便后续历史记录加载
            if conversation_id:
                placeholder_data = {
                    "file_type": "text",
                    "combined_content": combined_content[:500],  # 限制长度
                    "processing_method": "llm"
                }
                
                # 为文本内容生成嵌入向量
                embedding_texts = [f"Text content: {combined_content[:500]}"]
                ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
            
            return pure_llm_generate_answer(question, combined_content, temperature, max_tokens)
        elif route == "vlm":
            # VLM处理 - 优先处理图片文件，但需要考虑其他文件
            # 为了更好地处理混合内容，我们先处理图片，然后将其他文件内容作为上下文
            image_files = []
            table_content = ""  # 存储表格转换后的JSON内容
            other_content = ""  # 存储其他文件内容
            
            for file in files:
                file_path = file.name
                ext = os.path.splitext(file_path)[1].lower()
                
                if ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp"]:
                    image_files.append(file_path)
                elif ext in [".xlsx", ".xls", ".csv", ".docx", ".doc"]:  # 表格文件
                    try:
                        # 将表格文件转换为HOTree JSON格式
                        wrapped_file = types.SimpleNamespace(name=file.name)
                        tree_json = process_table_for_tree(wrapped_file, conversation_id=conversation_id)
                        if tree_json:
                            table_content += f"\n--- 表格文件 {os.path.basename(file_path)} 的HOTree JSON结构: {json.dumps(tree_json, ensure_ascii=False, indent=2)} ---\n"
                        else:
                            # 如果转换失败，尝试作为普通文本读取
                            try:
                                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                                    content = f.read(1000)  # 读取前1000个字符作为上下文
                                    other_content += f"\n--- 其他文件 {os.path.basename(file_path)} 内容: {content} ---\n"
                            except:
                                other_content += f"\n--- 其他文件 {os.path.basename(file_path)} (非文本文件) ---\n"
                    except Exception as e:
                        # 如果表格转换失败，尝试作为普通文本读取
                        try:
                            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                                content = f.read(1000)  # 读取前1000个字符作为上下文
                                other_content += f"\n--- 其他文件 {os.path.basename(file_path)} 内容: {content} ---\n"
                        except:
                            other_content += f"\n--- 其他文件 {os.path.basename(file_path)} (非文本文件) ---\n"
                else:  # 其他类型的文件
                    try:
                        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                            content = f.read(1000)  # 读取前1000个字符作为上下文
                            other_content += f"\n--- 其他文件 {os.path.basename(file_path)} 内容: {content} ---\n"
                    except:
                        other_content += f"\n--- 其他文件 {os.path.basename(file_path)} (非文本文件) ---\n"
            
            # 如果有图片文件，使用第一个图片文件并附加上下文
            if image_files:
                combined_context = ""
                if table_content:
                    combined_context += f"表格数据: {table_content}\n"
                if other_content:
                    combined_context += f"其他文件信息: {other_content}"
                
                # 为VLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "image",
                        "image_files": image_files,
                        "table_content": table_content,
                        "other_content": other_content,
                        "processing_method": "vlm"
                    }
                    
                    # 为图像上下文生成简单的嵌入向量
                    context_text = f"Image files: {', '.join([os.path.basename(img) for img in image_files])}"
                    if table_content:
                        context_text += f"; Table content: {table_content[:200]}"  # 限制长度
                    if other_content:
                        context_text += f"; Other content: {other_content[:200]}"  # 限制长度
                    
                    embedding_texts = [context_text]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                if combined_context:
                    enhanced_question = f"{question}\n\n{combined_context}"
                    return pure_vlm_generate_answer(enhanced_question, image_files[0], temperature, max_tokens)
                else:
                    return pure_vlm_generate_answer(question, image_files[0], temperature, max_tokens)
            else:
                # 如果没有找到图片文件但路线是VLM，使用第一个文件
                # 为VLM对话创建必要的处理文件，以便后续历史记录加载
                if conversation_id:
                    placeholder_data = {
                        "file_type": "image",
                        "file_path": files[0].name,
                        "processing_method": "vlm"
                    }
                    
                    # 为图像上下文生成简单的嵌入向量
                    embedding_texts = [f"Image file: {os.path.basename(files[0].name)}"]
                    ensure_conversation_cache(conversation_id, placeholder_data, embedding_texts)
                
                return pure_vlm_generate_answer(question, files[0].name, temperature, max_tokens)
        elif route == "hotree":
            # H-OTree处理 - 将所有表格文件合并为一棵树
            # 过滤出所有表格文件
            table_files = []
            for file in files:
                file_path = file.name if hasattr(file, 'name') else str(file)
                ext = os.path.splitext(file_path)[1].lower()
                if ext in [".xlsx", ".xls", ".csv", ".docx", ".doc"]:
                    table_files.append(file)
            
            if table_files:
                # --- 简化逻辑：只要有 pkl 就不重新解析 ---
                # 如果提供了 conversation_id，则使用历史记录文件夹，否则使用 cache 目录
                if conversation_id:
                    cache_dir = os.path.join("history", conversation_id)
                else:
                    cache_dir = "cache"
                cache_pkl = os.path.join(cache_dir, "temp.pkl")
                
                if os.path.exists(cache_pkl):
                    logger.info("检测到本地 H-OTree 缓存，跳过解析流程。")
                else:
                    logger.info("未检测到缓存，开始执行 H-OTree 完整解析...")
                    # 只有在没有缓存时才进行解析
                    data = process_multiple_tables_for_tree(table_files, conversation_id=conversation_id)
                    if not data:
                        return "多文件 H-OTree 解析失败"

                # 使用H-OTree方法回答问题
                result = process_question_only(question, temperature, max_tokens, conversation_id=conversation_id)
                return result
        else:
            return f"未知处理线路: {route}"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ 多文件处理失败: {error_msg}")
        return "多文件处理失败"


def process_question_only(question, temperature=0.5, max_tokens=2048, conversation_id=None):
    """专门处理问题，返回答案"""
    # 如果提供了 conversation_id，则使用历史记录文件夹中的文件，否则使用默认路径
    if conversation_id:
        cache_dir = os.path.join("history", conversation_id)
        
        # 检查缓存目录中的必要文件是否存在（特别是当conversation_id存在时，即从历史记录加载）
        required_files = [
            os.path.join(cache_dir, "temp.pkl"),
            os.path.join(cache_dir, "temp.embedding.json")
        ]
        
        missing_files = [f for f in required_files if not os.path.exists(f)]
        
        if missing_files:
            # 如果缺少必要的缓存文件，说明这个历史记录无法继续处理
            gr.Warning(f"无法继续此对话：缺少必要的处理文件。请重新上传文件后再继续。")
            return "无法继续此对话：缺少必要的处理文件。请重新上传文件后再继续。"
        
        # 读取占位符数据以确定对话类型
        pkl_file = os.path.join(cache_dir, "temp.pkl")
        with open(pkl_file, 'rb') as f:
            placeholder_data = pickle.load(f)
        
        # 根据对话类型选择处理方法
        if isinstance(placeholder_data, dict):
            processing_method = placeholder_data.get("processing_method", "hotree")
            file_type = placeholder_data.get("file_type", "table")
        else:
            # 旧版本数据格式，假设为表格类型
            processing_method = "hotree"
            file_type = "table"
        
        # 根据处理方法决定如何处理问题
        if processing_method == "vlm":
            # 对于VLM对话，使用VLM方法继续处理
            image_files = placeholder_data.get("image_files", [])
            file_path = placeholder_data.get("file_path", "")
            table_content = placeholder_data.get("table_content", "")
            other_content = placeholder_data.get("other_content", "")
            
            # 构建上下文
            combined_context = ""
            if table_content:
                combined_context += f"表格数据: {table_content}\n"
            if other_content:
                combined_context += f"其他文件信息: {other_content}"
            
            # 选择要使用的图片文件
            image_to_use = image_files[0] if image_files else file_path
            if image_to_use:
                enhanced_question = question
                if combined_context:
                    enhanced_question = f"{question}\n\n{combined_context}"
                
                return pure_vlm_generate_answer(enhanced_question, image_to_use, temperature, max_tokens)
            else:
                # 如果没有图片文件，返回错误
                gr.Warning("VLM对话缺少图片文件")
                return "VLM对话缺少图片文件"
                
        elif processing_method == "llm":
            # 对于LLM对话，使用LLM方法继续处理
            combined_content = placeholder_data.get("combined_content", "")
            file_content_preview = placeholder_data.get("file_content_preview", "")
            
            # 合并所有可用内容
            full_content = combined_content or file_content_preview
            if full_content:
                return pure_llm_generate_answer(question, full_content, temperature, max_tokens)
            else:
                # 如果没有内容，可以使用通用LLM回答
                return pure_llm_generate_answer(question, "", temperature, max_tokens)
        
        else:
            # 对于HOTree对话，使用原来的处理方法
            table_file = os.path.join("history", conversation_id, "temp.xlsx")
            if not os.path.exists(table_file):
                gr.Warning("表格文件丢失，无法继续处理此对话")
                return "表格文件丢失，无法继续处理此对话"
    else:
        table_file = "data/SSTQA/temp_tables/temp.xlsx"
        cache_dir = "cache"
        
        if not os.path.exists(table_file):
            gr.Warning("Please upload a table first")
            return "Please upload a table first"
    
    if not question.strip():
        gr.Warning("Please enter a question")
        return "Please enter a question"
    
    # 对于HOTree对话（包括没有conversation_id的新对话），使用原来的处理方法
    try:
        # 记录参数变更日志（使用 loguru 格式：时间 | 级别 | 内容）
        param_log_file = os.path.join(LOG_DIR, "param_change.log")
        os.makedirs(LOG_DIR, exist_ok=True)
        from datetime import datetime
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
        msg = f"{timestamp} | PARAM_CHANGE | temperature={temperature}, max_tokens={max_tokens}\n"
        with open(param_log_file, "a", encoding="utf-8") as f:
            f.write(msg)
        qa_pair = {
            "id": "temp",
            "table_id": "temp",
            "query": question.strip()
        }
        result=answer_question(
            qa_pair=qa_pair,
            table_file=table_file,
            cache_dir=cache_dir,
            enable_emebdding=True,
            enable_query_decompose=True,
            log_dir=LOG_DIR,
            temperature=temperature,
            max_tokens=max_tokens
        )
        if result :
            gr.Info("✅ Answer generated successfully!")
            return f"Answer: {result.get('model_output', 'No answer')}\n\nConfidence: {result.get('reliability', 'Unknown')}"
        else:
            gr.Warning("❌ Failed to generate answer")
            return "Failed to generate answer"
    except Exception as e:
        import traceback
        error_msg = f"处理错误: {str(e)}\n错误详情: {traceback.format_exc()}"
        gr.Warning(f"❌ Failed to generate answer: {error_msg}")
        return "Failed to generate answer"

def clear_all():
    """清除所有内容并删除相关文件"""
    import shutil
    import os    
    # 返回空输出以匹配界面清空
    empty_outputs = (None, "", "", {}, None, "")
    # 删除临时表格文件
    temp_dir = "data/SSTQA/temp_tables"
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
        os.makedirs(temp_dir, exist_ok=True)  # 重新创建空目录
    
    # 删除log目录下的所有文件
    log_dir = LOG_DIR
    clear_directory_contents(log_dir)

    # 删除cache目录下的所有文件
    cache_dir = "cache"
    clear_directory_contents(cache_dir)
    
    # 重置思维链条数据
    global thinking_chain_data
    thinking_chain_data = {
        "question_answering": {},
        "retrieval_chains": []
    }
    
    return None, "", "", {}, None, ""  # 清空所有界面组件（含图谱）

def get_thinking_chain():
    """获取思维链条数据"""
    global thinking_chain_data
    return thinking_chain_data

def read_all_logs(log_dir=LOG_DIR, max_lines=200):
    """合并读取所有日志文件，按时间顺序显示，并添加颜色美化"""
    all_lines = []
    
    # 优先读取的特定日志文件
    priority_log_files = [
        os.path.join(log_dir, "temp.xlsx.log"),
        os.path.join(log_dir, "param_change.log"),
        os.path.join(log_dir, "temp.log"),
    ]
    
    # 读取所有 .log 文件（包括app.log）
    all_log_files = []
    if os.path.exists(log_dir):
        for file in os.listdir(log_dir):
            if file.endswith('.log'):
                file_path = os.path.join(log_dir, file)
                if os.path.isfile(file_path):
                    all_log_files.append(file_path)
    
    # 合并优先文件和所有日志文件，去重
    log_files = list(set(priority_log_files + all_log_files))
    
    for log_path in log_files:
        if os.path.exists(log_path):
            try:
                with open(log_path, "r", encoding="utf-8", errors="ignore") as f:
                    lines = f.readlines()
                    all_lines.extend(lines)
            except Exception as e:
                all_lines.append(f"[ERROR] 读取 {log_path} 失败: {e}\n")
    
    # 按时间戳排序（loguru 格式：时间 | 级别 | ...）
    try:
        all_lines.sort(key=lambda x: x.split("|")[0].strip() if "|" in x else "")
    except Exception:
        pass
    
    # 取最后 max_lines 行
    log_content = "".join(all_lines[-max_lines:]) if all_lines else "暂无日志"
    
    # 添加颜色美化 - 将日志转换为HTML格式
    # 支持 loguru 格式: 时间 | 级别 | 内容
    html_lines = []
    for line in log_content.split("\n"):
        if "|" in line and len(line.split("|")) >= 3:
            parts = line.split("|", 2)
            timestamp = parts[0].strip()
            level = parts[1].strip()
            content = parts[2].strip()
            # 为时间戳添加蓝色，为日志级别添加绿色
            html_line = f"<span style='color: blue'>{timestamp}</span> | <span style='color: green'>{level}</span> | {content}<br>"
        else:
            # 非标准格式行保持原样
            html_line = line + "<br>"
        html_lines.append(html_line)
    
    # 包装在<pre>标签中以保留格式，但使用HTML允许颜色显示
    return f"<pre style='font-family: monospace; white-space: pre-wrap; word-wrap: break-word;'>{' '.join(html_lines)}</pre>"


def tree_json_to_table_dict(tree_json, table_name="edited_table"):
    """
    将前端树 JSON（列表）递归转换为 construct_feature_tree 可用的 table_dict:
    {table_name: <嵌套结构/list/值>}#处理三种情况 dict of dict /dict with list /listofdict 无listoflist

    规则：
     他的规则是这样的：读json 读到每一个node
     1.如果有children而且有很多个children 构建一个dict dict的name是dict的index 然后dict的children是他的value value是一个list，返回这个dict
     2.如果只有一个children 那么他的value就是个string/float 总而言之就是一个值，就构建一个dict index=项目 value=children 
     3.如果没有children 就返回name（string类型）不建立dict
    - 
    """
    def convert_node(node: dict):   #处理dict with list of dict  返回一个dict
                                    
        if not isinstance(node, dict):
            return None
        name = str(node.get("name", ""))
        children = node.get("children", [])
        if children:
            if len(children)==1:
                child=children[0]
                return {name:convert_node(child)} #返回一个dict 
            # 孩子是一个list 直接传入convert children
            else:
                merged=convert_children(children)
                return {name: merged}#name+list
        # 无 children -> 叶子，返回 name 字符串
        return name

    def convert_children(children: list):#输入一个list 如果list里面每一个dict名字都是序号 返回一个list 里面是去掉序号包装的dict
                                         #如果不是这样 那么返回一个dict
        if not children:#
            return []
        else:#如果所有的孩子都是name：【n】，那么处理每个孩子的孩子为一个dict，加到list 返回list
            if all(isinstance(c, dict) and re.fullmatch(r"\[\d+\]", str(c.get("name", ""))) for c in children):
                lst = []
                for c in children:
                    v = c.get("children", [])
                    if isinstance(v, list):#一般来说序号的value就是list
                        merged = {}
                        for item in v:
                            item1=convert_node(item);
                            if isinstance(item1, dict) and len(item1) == 1:
                                k, val = next(iter(item1.items()))
                                merged[k] = val#组装成一个dict
                            else:#防御性 应该不会
                                merged[str(len(merged))] = item1
                        lst.append(merged)#加到list中
                return lst#返回不带序号的list

        # 保持顺序的列表，不再合并/排序
        #如果孩子的名字都是正常的那么这个时候传入的
        return [convert_node(c) for c in children]

    if not isinstance(tree_json, list) or not tree_json:
        return {table_name: []}

    # 使用最外层节点作为表节点，不再下钻 children[0]
    #tree_json是一个list of dict 暂时认为只有一个dict
    outer = tree_json[0] if isinstance(tree_json[0], dict) else {}
    logger.info(f"[debug] parsed root: {outer}")
    table_node = outer
    table_name = table_node.get("name", table_name) if isinstance(table_node, dict) else table_name
    table_children = table_node.get("children", []) if isinstance(table_node, dict) else []

    logger.info(f"[debug] table_name: {table_name}, child_count: {len(table_children)}")
    table_body = convert_children(table_children)#直接传入children_list
    return {table_name: table_body}#返回一个dict


def construct_feature_tree_simple(obj, name="root"):#认为一定是dict+list of dict的结构，没有listof list
    
    index_tree = IndexTree()
    body_tree = BodyTree()

    def add_pair(idx_value, body_value):
        index_node = IndexNode(value=idx_value)
        body_node = BodyNode(body_value)
        index_node.body = [body_node]
        index_tree.add_index(index_node)
        body_tree.add_deep(body_node)

    if isinstance(obj, dict):
        for k, v in obj.items():
            logger.info(f"[debug] dict key: {k}")
            if isinstance(v, (dict, list)):#如果value是list
                child_tree = construct_feature_tree_simple(v, name=str(k))
                add_pair(str(k), child_tree)
            else:
                add_pair(str(k), v)#如果value是string/float 那么直接添加到body tree中
        return FeatureTree(index_tree=index_tree, body_tree=body_tree)

    if isinstance(obj, list):  # 行列表或通用列表
        for i, v in enumerate(obj):
            idx = f"[{i}]"
            if isinstance(v, (dict, list)):
                child_tree = construct_feature_tree_simple(v, name=idx)
                add_pair(idx, child_tree)
            else:
                add_pair(idx, v)
        return FeatureTree(index_tree=index_tree, body_tree=body_tree)

    # 原子值兜底
    add_pair(str(name) if name is not None else "root", obj)
    return FeatureTree(index_tree=index_tree, body_tree=body_tree)


def rebuild_feature_tree_from_json(tree_json, cache_dir="cache", temp_dir="data/SSTQA/temp_tables", log_dir=LOG_DIR):
    try:
        def strip_ids(obj):
            if isinstance(obj, list):
                return [strip_ids(o) for o in obj]
            if isinstance(obj, dict):
                return {
                    k: strip_ids(v)
                    for k, v in obj.items()
                    if k != "id"
                }
            return obj

        # 临时调试日志，观测前端传入结构（仅取前 1 个元素，防止日志过大）
        try:
            sample = tree_json[:1] if isinstance(tree_json, list) else tree_json
            logger.info(f"[debug] tree_json sample: {sample}")
        except Exception as e:
            logger.warning(f"[debug] cannot log tree_json sample: {e}")

        os.makedirs(cache_dir, exist_ok=True)
        os.makedirs(temp_dir, exist_ok=True)
        os.makedirs(log_dir, exist_ok=True)

        # 去掉 id，防止写回文件
        cleaned_tree = strip_ids(tree_json)

        # 持久化最新 JSON（输入版本）
        with open(os.path.join(cache_dir, "temp.json"), "w", encoding="utf-8") as f:
            json.dump(cleaned_tree, f, ensure_ascii=False, indent=2)
        with open(os.path.join(temp_dir, "temp.json"), "w", encoding="utf-8") as f:
            json.dump(cleaned_tree, f, ensure_ascii=False, indent=2)

        # JSON -> FeatureTree
        tree_dict = tree_json_to_table_dict(cleaned_tree)
        logger.info(f"[debug] parsed tree_dict: {tree_dict}")

        # 尝试用简化构造逻辑，以便保留 list 内的 dict 结构
        ho_tree = None
        try:
            if isinstance(tree_dict, dict) and len(tree_dict) == 1:
                t_name, t_body = next(iter(tree_dict.items()))
                ho_tree = construct_feature_tree_simple({t_name: t_body}, name="root")
                logger.info("[debug] ho_tree built by construct_feature_tree_simple")
            else:
                ho_tree = construct_feature_tree_simple(tree_dict, name="root")
        except Exception as ee:
            logger.warning(f"[debug] simple construct failed, fallback to original: {ee}")
            ho_tree = construct_feature_tree(tree_dict)

        try:
            idx_children = [n.value for n in ho_tree.index_tree.root.children]
            logger.info(f"[debug] ho_tree index root children: {idx_children}")
        except Exception as ee:
            logger.warning(f"[debug] cannot log index tree children: {ee}")

        ho_tree = tag_feature_tree(ho_tree)

        # 保存文本、对象、embedding
        with open(os.path.join(cache_dir, "temp.txt"), "w", encoding="utf-8") as f:
            f.write(ho_tree.__str__([1]))
        with open(os.path.join(cache_dir, "temp.pkl"), "wb") as f:
            pickle.dump(ho_tree, f)
        # 额外保存 HO-Tree 规范化 JSON，若原始仅单表，则包上一层表名，避免 __json__ 展开丢失表头
        ho_json = ho_tree.__json__()
        with open(os.path.join(cache_dir, "temp.ho.json"), "w", encoding="utf-8") as f:
            json.dump(ho_json, f, ensure_ascii=False, indent=4)

        # 生成 embedding，失败不阻塞（确保输入为字符串列表）
        try:
            raw_values = ho_tree.all_value_list()
            texts = [str(x) for x in raw_values] if raw_values else []
            if texts:
                embedding_dict = EmbeddingModel().get_embedding_dict(texts)
                EmbeddingModel().save_embedding_dict(
                    embedding_dict,
                    os.path.join(cache_dir, "temp.embedding.json")
                )
            else:
                logger.warning("embedding skipped: empty value list")
        except Exception as ee:
            logger.error(f"embedding generate failed: {ee}")

        return True, "ok"
    except Exception as e:
        logger.error(f"rebuild_feature_tree_from_json failed: {e}")
        return False, str(e)


def save_conversation_history(conversation_id, chat_history):
    """
    保存对话历史到文件
    
    参数:
    conversation_id: 对话ID
    chat_history: 对话历史，格式为[{"role": "user", "content": "问题"}, {"role": "assistant", "content": "回答"}, ...]
    """
    if not conversation_id:
        print(f"[DEBUG] save_conversation_history: conversation_id 为空")
        return False
    
    try:
        # 获取当前工作目录和绝对路径
        current_dir = os.getcwd()
        print(f"[DEBUG] save_conversation_history: 当前工作目录: {current_dir}")
        
        # 确保历史目录存在
        history_dir = os.path.join("history", conversation_id)
        history_dir_abs = os.path.abspath(history_dir)
        print(f"[DEBUG] save_conversation_history: 历史目录(相对): {history_dir}")
        print(f"[DEBUG] save_conversation_history: 历史目录(绝对): {history_dir_abs}")
        
        os.makedirs(history_dir, exist_ok=True)
        print(f"[DEBUG] save_conversation_history: 目录创建/存在: {os.path.exists(history_dir)}")
        
        # 保存对话历史到文件
        history_file = os.path.join(history_dir, "chat_history.json")
        history_file_abs = os.path.abspath(history_file)
        print(f"[DEBUG] save_conversation_history: 保存到文件(相对): {history_file}")
        print(f"[DEBUG] save_conversation_history: 保存到文件(绝对): {history_file_abs}")
        print(f"[DEBUG] save_conversation_history: 消息数量 {len(chat_history)}")
        print(f"[DEBUG] save_conversation_history: 消息内容 {chat_history}")
        
        with open(history_file, 'w', encoding='utf-8') as f:
            json.dump(chat_history, f, ensure_ascii=False, indent=2)
        
        # 验证文件是否真的被写入
        file_exists = os.path.exists(history_file)
        file_size = os.path.getsize(history_file) if file_exists else 0
        print(f"[DEBUG] save_conversation_history: 文件写入后存在: {file_exists}, 大小: {file_size} 字节")
        
        if not file_exists:
            print(f"[DEBUG] save_conversation_history: 警告！文件写入后不存在！")
            return False
        
        print(f"[DEBUG] save_conversation_history: 保存成功")
        return True
    except Exception as e:
        print(f"[DEBUG] save_conversation_history: 保存失败 - {e}")
        import traceback
        traceback.print_exc()
        logger.error(f"保存对话历史失败: {e}")
        return False


def load_conversation_history(conversation_id):
    """
    从文件加载对话历史
    
    参数:
    conversation_id: 对话ID
    
    返回:
    对话历史，格式为[{"role": "user", "content": "问题"}, {"role": "assistant", "content": "回答"}, ...]
    """
    if not conversation_id:
        print(f"[DEBUG] load_conversation_history: conversation_id 为空")
        return []
    
    try:
        # 获取当前工作目录和绝对路径
        current_dir = os.getcwd()
        print(f"[DEBUG] load_conversation_history: 当前工作目录: {current_dir}")
        
        history_file = os.path.join("history", conversation_id, "chat_history.json")
        history_file_abs = os.path.abspath(history_file)
        print(f"[DEBUG] load_conversation_history: 加载文件(相对): {history_file}")
        print(f"[DEBUG] load_conversation_history: 加载文件(绝对): {history_file_abs}")
        
        file_exists = os.path.exists(history_file)
        print(f"[DEBUG] load_conversation_history: 文件存在(相对路径): {file_exists}")
        
        # 也尝试绝对路径
        file_exists_abs = os.path.exists(history_file_abs)
        print(f"[DEBUG] load_conversation_history: 文件存在(绝对路径): {file_exists_abs}")
        
        # 列出目录内容，看看实际有什么文件
        history_dir = os.path.join("history", conversation_id)
        if os.path.exists(history_dir):
            dir_contents = os.listdir(history_dir)
            print(f"[DEBUG] load_conversation_history: 目录内容: {dir_contents}")
        
        if file_exists:
            with open(history_file, 'r', encoding='utf-8') as f:
                chat_history = json.load(f)
            print(f"[DEBUG] load_conversation_history: 加载成功，消息数量: {len(chat_history) if isinstance(chat_history, list) else 0}")
            print(f"[DEBUG] load_conversation_history: 消息内容: {chat_history}")
            return chat_history if isinstance(chat_history, list) else []
        else:
            print(f"[DEBUG] load_conversation_history: 文件不存在")
            return []
    except Exception as e:
        print(f"[DEBUG] load_conversation_history: 加载失败 - {e}")
        import traceback
        traceback.print_exc()
        logger.error(f"加载对话历史失败: {e}")
        return []
