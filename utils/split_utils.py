import openpyxl
from openpyxl.utils import range_boundaries

from utils.sheet_utils import *


# TODO
def build_rowspan_entire_dict_recur(sheet, merged_cell_list):
    conti_start = []    # 记录连续整行合并的开始位置x
    index = 0
    while index < len(merged_cell_list) - 1:
        # 判断是否连续
        x1, y1, x2, y2 = get_merge_cell_size(sheet, merged_cell_list[index].coordinate)
        xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, merged_cell_list[index + 1].coordinate)
        
        if x2 + 1 == xx1:   # 发现整行连续
            conti_start.append()
        
        
        # 跳过该部分连续
        
        index += 1
    
    
    if len(conti_start) == 0:   # 不存在连续的整行合并
        pass
    elif len(conti_start) == 1:    # 存在一次连续的整行合并
        pass
    else:   # 存在多次连续的整行合并
        pass
    
    return None


def rowspan_entire(sheet):
    """处理横跨整张表的合并行"""

    nrows = sheet.max_row
    ncols = sheet.max_column

    # Step1 找到所有横跨整张表的合并单元格位置
    merged_cells = sheet.merged_cells
    merged_cells = sorted(
        merged_cells, key=lambda x: (x.min_row, x.min_col)
    )  # 将合并单元格位置排序

    merged_cell_list = []
    for cell in merged_cells:
        y1, x1, y2, x2 = cell.bounds
        if y1 == 1 and y2 == ncols:
            # Step2 获取key以及subsheet
            merged_cell_list.append(cell)

    # 没有横跨整张表的合并单元格，这一模式无匹配结果，默认表格名称为table
    if len(merged_cell_list) == 0 or nrows == 1 or ncols == 1:
        return {DEFAULT_TABLE_NAME: sheet}

    # TODO
    # res = build_rowspan_entire_dict_recur(sheet, merged_cell_list)

    res = {}
    for index, cell in enumerate(merged_cell_list):
        y1, x1, y2, x2 = cell.bounds
        key = sheet.cell(row=x1, column=1).value
        if index == len(merged_cell_list) - 1:
            if x2 == nrows:  # 最后一个整行的合并单元格即为表格最后一行
                res[key] = None
            else:  # 最后一个整行的合并单元格不是表格最后一行
                res[key] = get_sub_sheet(sheet, x2 + 1, 1, nrows, ncols).active
        else:
            res[key] = get_sub_sheet(
                sheet, x2 + 1, 1, merged_cell_list[index + 1].min_row - 1, ncols
            ).active
    return res


def colspan_entire(sheet):
    """处理纵跨整张表的合并列"""
    nrows = sheet.max_row
    ncols = sheet.max_column

    # Step1 找到所有纵跨整张表的合并单元格位置
    merged_cells = sheet.merged_cells
    merged_cells = sorted(
        merged_cells, key=lambda x: (x.min_row, x.min_col)
    )  # 将合并单元格位置排序

    merged_cell_list = []
    for cell in merged_cells:
        y1, x1, y2, x2 = cell.bounds
        if x1 == 1 and x2 == nrows:
            # Step2 获取key以及subsheet
            merged_cell_list.append(cell)

    # 没有横跨整张表的合并单元格，这一模式无匹配结果，默认表格名称为table
    if len(merged_cell_list) == 0 or nrows == 1 or ncols == 1:
        return {DEFAULT_TABLE_NAME: sheet}

    res = {}
    for index, cell in enumerate(merged_cell_list):
        y1, x1, y2, x2 = cell.bounds
        key = sheet.cell(row=1, column=y1).value
        if index == len(merged_cell_list) - 1:
            if y2 == nrows:  # 最后一个整列的合并单元格即为表格最后一列
                res[key] = None
            else:  # 最后一个整列的合并单元格不是表格最后一列
                res[key] = get_sub_sheet(sheet, 1, y2 + 1, nrows, ncols).active
        else:
            res[key] = get_sub_sheet(
                sheet, 1, y2 + 1, nrows, merged_cell_list[index + 1].min_col - 1
            ).active
    return res


def split_subtable_row(sheet):
    """将表按行拆为并列的几个部分"""
    nrows = sheet.max_row  # 高度
    ncols = sheet.max_column  # 长度
    # Step1 保证粒度向下递减，在第一个不递减的地方停止并记录位置（用于分割）
    split_pos_list = [1]
    for row in range(1, nrows):
        col = 1
        while col <= ncols:

            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            width = y2 - y1 + 1

            n_cell = sheet.cell(row=row + 1, column=col)
            n_x1, n_y1, n_x2, n_y2 = get_merge_cell_size(sheet, n_cell.coordinate)
            n_width = n_y2 - n_y1 + 1

            if n_width > width:
                split_pos_list.append(n_x1)
                break
            col += 1
    # Step2 按照分割的位置，获取表格的每一子部分
    if len(split_pos_list) == 1:
        # 没有需要分割的子部分，即粒度从上到下是单调递减的，返回None，这一步匹配无结果
        return [sheet]
    wb_list = []
    for index, start_row in enumerate(split_pos_list):
        if index == len(split_pos_list) - 1:
            wb_list.append(get_sub_sheet(sheet, start_row, 1, nrows, ncols).active)
        else:
            wb_list.append(
                get_sub_sheet(
                    sheet, start_row, 1, split_pos_list[index + 1] - 1, ncols
                ).active
            )

    return wb_list


def get_row_cell_number(sheet, row):
    n = 0

    col = 1
    while col <= sheet.max_column:
        cell = sheet.cell(row=row, column=col)
        x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
        col += y2 - y1 + 1
        n += 1

    return n


def split_subtable_each_row(sheet, schema_height):
    """每一行都拆分为一个子部分"""
    nrows = sheet.max_row  # 高度
    ncols = sheet.max_column  # 长度

    sheet_dict = {}

    row = 1
    while row <= nrows:
        cell = sheet.cell(row=row, column=schema_height)
        x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)

        key = str(get_merge_cell_value(sheet, cell.coordinate))
        if y1 != 1:  # 获得嵌套 Schema序列
            y = y1 - 1
            while y >= 1:
                ccell = sheet.cell(row=row, column=y)
                xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, ccell.coordinate)
                key = (
                    str(get_merge_cell_value(sheet, ccell.coordinate))
                    + DEFAULT_SPLIT_SIG
                    + key
                )
                y -= yy2 - yy1 + 1

        n_row_cells = get_row_cell_number(sheet, row)
        if n_row_cells % 2 == 0:  # 这一行有双数个单元格，说明是kv pair的拼接
            xx1, yy1, xx2, yy2 = get_merge_cell_size(
                sheet, sheet.cell(row=row, column=schema_height + 1).coordinate
            )
            if yy2 == ncols:
                sheet_dict[key] = get_merge_cell_value(
                    sheet, sheet.cell(row=row, column=schema_height + 1).coordinate
                )
            else:
                index = 0
                while f"{DEFAULT_SUBTABLE_NAME}{index}" in sheet_dict:
                    index += 1
                sheet_dict[f"{DEFAULT_SUBTABLE_NAME}{index}"] = get_sub_sheet(
                    sheet, x1, 1, x2, ncols
                ).active
        else:  # 这一行有单数个单元格，说明kv pair list是value，外面还有一个key
            sheet_dict[key] = get_sub_sheet(sheet, x1, y2 + 1, x2, ncols).active

        row += x2 - x1 + 1

    return sheet_dict


def split_subtable_column(sheet):
    """将表按行拆为并列的几个部分"""
    nrows = sheet.max_row  # 高度
    ncols = sheet.max_column  # 长度
    # Step1 保证粒度向下递减，在第一个不递减的地方停止并记录位置（用于分割）
    split_pos_list = [1]
    for row in range(1, nrows):
        col = 1
        while col <= ncols:

            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            width = y2 - y1 + 1

            n_cell = sheet.cell(row=row + 1, column=col)
            n_x1, n_y1, n_x2, n_y2 = get_merge_cell_size(sheet, n_cell.coordinate)
            n_width = n_y2 - n_y1 + 1

            if n_width != width:
                split_pos_list.append(n_x1)
                break
            col += 1
    # Step2 按照分割的位置，获取表格的每一子部分
    if len(split_pos_list) == 1:
        # 没有需要分割的子部分，即粒度从上到下是单调递减的，返回None，这一步匹配无结果
        return [sheet]
    wb_list = []
    for index, start_row in enumerate(split_pos_list):
        if index == len(split_pos_list) - 1:
            wb_list.append(get_sub_sheet(sheet, start_row, 1, nrows, ncols).active)
        else:
            wb_list.append(
                get_sub_sheet(
                    sheet, start_row, 1, split_pos_list[index + 1] - 1, ncols
                ).active
            )

    return wb_list


def min_match(sheet):
    """匹配最小结构: 左右/上下 返回一个字典 如果不是最小结构则返回None"""
    nrows = sheet.max_row
    ncols = sheet.max_column

    # 直接获得左上角和右下角的cell, 看是否是只有两个cell
    top_left_cell = sheet.cell(row=1, column=1)
    down_right_cell = sheet.cell(row=nrows, column=ncols)

    a1, b1, a2, b2 = get_merge_cell_size(sheet, top_left_cell.coordinate)
    x1, y1, x2, y2 = get_merge_cell_size(sheet, down_right_cell.coordinate)

    # 如果可以左右拼接或上下拼接则说明是最小单元，否则返回None
    if (a1 == x1 and a2 == x2 and b2 + 1 == y1) or (
        b1 == y1 and b2 == y2 and a2 + 1 == x1
    ):  # 左右 / 上下
        key = top_left_cell.value
        if x1 == x2 and y1 == y2:
            value = down_right_cell.value
        else:
            for merged_range in sheet.merged_cells.ranges:
                if down_right_cell.coordinate in merged_range:
                    value = sheet[merged_range.start_cell.coordinate].value
        return {key: value}
    return None


def extract_schema_row(sheet):
    """递归的提取横向排列的Schema, 返回嵌套的list形式"""
    nrows = sheet.max_row
    ncols = sheet.max_column

    # 按列循环, 提取每一大列
    schema = []
    col = 1
    while col <= ncols:
        cell = sheet.cell(row=1, column=col)
        x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)

        if not single_cell(sheet, x1, y1, nrows, y2):  # 该列还有嵌套的Schema
            sub_schema_wb = get_sub_sheet(sheet, x2 + 1, y1, nrows, y2)
            schema.append(
                {
                    sheet.cell(row=x1, column=y1).value: extract_schema_row(
                        sub_schema_wb.active
                    )
                }
            )
        else:
            schema.append(sheet.cell(row=x1, column=y1).value)

        col += y2 - y1 + 1
    return schema


def extract_schema_column(sheet):
    """递归的提取纵向排列的Schema, 返回嵌套的list形式"""
    nrows = sheet.max_row
    ncols = sheet.max_column

    # 按行循环, 提取每一大行
    schema = []
    row = 1
    while row <= nrows:
        cell = sheet.cell(row=row, column=1)
        x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)

        if not single_cell(sheet, x1, y1, x2, ncols):  # 该行还有嵌套的Schema
            sub_schema_wb = get_sub_sheet(sheet, x1, y2 + 1, x2, ncols)
            schema.append(
                {
                    sheet.cell(row=x1, column=y1).value: extract_schema_column(
                        sub_schema_wb.active
                    )
                }
            )
        else:
            schema.append(sheet.cell(row=x1, column=y1).value)

        row += x2 - x1 + 1
    return schema


def flatten_schema(schema):
    """将嵌套形式的Schema展平到最细粒度, 返回Schema列表"""
    flattenned_schema = []
    for attr in schema:
        if type(attr) is dict:
            key = next(iter(attr))
            tmp = flatten_schema(attr[key])
            for v in tmp:
                if type(v) is list:
                    flattenned_schema.append([key] + v)
                else:
                    flattenned_schema.append([key, v])
        else:
            flattenned_schema.append(attr)
    return flattenned_schema


"""
[['教育学院  （654人，其中师范类毕业生567人）', 
    [['本科', 
        [['小学教育', '本科', 25], 
         ['小学教育（公费）', '本科', 20], 
         ['学前教育', '本科', 177], 
         ['学前教育（专升本）', '本科', 151], 
         ['学前教育本（春季高考）', '本科', 40], 
         ['学前教育本（国际方向）', '本科', 58], 
         ['应用心理学(非师）', '本科', 40]
         ]
     ], 
     ['专科', 
        [['学前教育专（国际方向）', '专科', 96], 
        ['早期教育（非师）', '专科', 47]]
     ]
    ], 
    '宫老师', 
    [['0531-86526675'], ['0531-86526675']]
], 
['艺术设计学院\n（413人）', [['本科', [['产品设计', '本科', 38], ['服装与服饰设计', '本科', 48], ['服装与服饰设计（专升本）', '本科', 40], ['环境设计', '本科', 49], ['视觉传达设计', '本科', 46], ['视觉传达设计（专升本）', '本科', 45], ['数字媒体艺术', '本科', 31], ['数字媒体艺术（校企合作）', '本科', 28], ['数字媒体艺术（专升本）', '本科', 49]]], ['专科', [['人物形象设计', '专科', 38], ['环境艺术设计', '专科', 1]]]], '宫老师', '0531-86526671']]
"""


def schema_content_match(flattenned_schema, row_content):
    """给定平铺的Schema和行列表, 返回嵌套的字典表格结构

    Args:
        flattenned_schema (_type_): 平铺的Schema列表
        row_content (_type_): 嵌套的字典表格结构

    Returns:
        _type_: _description_
    """

    sub_value_index = 0
    return_list = []
    for row in row_content:
        key_index = 0
        return_dict = {}
        for col_index, value in enumerate(row):
            if type(value) == list and len(value) > 1:
                return_dict[f"subvalue_{sub_value_index}"], key_add = (
                    schema_content_match(flattenned_schema[key_index:], value)
                )
                sub_value_index += 1
                key_index += key_add
            else:
                if type(value) == list and len(value) == 1:
                    value = value[0]
                if type(flattenned_schema[key_index]) == list:
                    return_dict = merge_json(
                        return_dict,
                        build_nested_dict(flattenned_schema[key_index], value),
                    )
                else:
                    return_dict[flattenned_schema[key_index]] = value
                key_index += 1
        return_list.append(return_dict)
    return return_list, key_index


def extract_rows(sheet):
    nrows = sheet.max_row
    ncols = sheet.max_column

    row_content = []
    row = 1
    while row <= nrows:
        col = 1
        x1, y1, x2, y2 = get_merge_cell_size(
            sheet, sheet.cell(row=row, column=1).coordinate
        )
        height = x2 - x1 + 1
        one_row = []
        while col <= ncols:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            if single_cell(sheet, x1, y1, x1 + height - 1, y2):
                one_row.append(sheet.cell(row=x1, column=y1).value)
                col += y2 - y1 + 1
            else:
                # 找到粒度变化反向的位置
                pos = ncols + 1
                for col_i in range(col + 1, ncols + 1):
                    row_i = row
                    while row_i <= x1 + height - 1:
                        cell = sheet.cell(row=row_i, column=col_i)
                        xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, cell.coordinate)
                        hheight = xx2 - xx1 + 1

                        n_cell = sheet.cell(row=row_i, column=col_i + 1)
                        n_xx1, n_yy1, n_xx2, n_yy2 = get_merge_cell_size(
                            sheet, n_cell.coordinate
                        )
                        n_hheight = n_xx2 - n_xx1 + 1

                        if n_hheight > hheight:
                            pos = n_yy1
                            break
                        row_i += 1
                one_row.append(
                    extract_rows(
                        get_sub_sheet(sheet, x1, y1, x1 + height - 1, pos - 1).active
                    )
                )
                col = pos

        row_content.append(one_row)
        row += height

    return row_content


def extract_columns(sheet):
    nrows = sheet.max_row
    ncols = sheet.max_column

    col_content = []
    col = 1
    while col <= ncols:
        row = 1
        x1, y1, x2, y2 = get_merge_cell_size(
            sheet, sheet.cell(row=1, column=col).coordinate
        )
        width = y2 - y1 + 1
        one_col = []
        while row <= nrows:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            if single_cell(sheet, x1, y1, x2, y1 + width - 1):
                one_col.append(sheet.cell(row=x1, column=y1).value)
                row += x2 - x1 + 1
            else:
                # 找到粒度变化反向的位置
                pos = nrows + 1
                for row_i in range(1, nrows + 1):
                    col_i = col
                    while col_i <= x1 + width - 1:
                        cell = sheet.cell(row=row_i, column=col_i)
                        xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, cell.coordinate)
                        wwidth = xx2 - xx1 + 1

                        n_cell = sheet.cell(row=row_i + 1, column=col_i)
                        n_xx1, n_yy1, n_xx2, n_yy2 = get_merge_cell_size(
                            sheet, n_cell.coordinate
                        )
                        n_wwidth = n_yy2 - n_yy1 + 1

                        if n_wwidth > wwidth:
                            pos = n_xx1
                            break
                        col_i += 1
                one_col.append(
                    extract_columns(
                        get_sub_sheet(sheet, x1, y1, pos - 1, y1 + width - 1).active
                    )
                )
                row = pos

        col_content.append(one_col)
        col += width

    return col_content


def split_schema_row(sheet):
    """给定sheet，返回schema和剩下的数据"""
    nrows = sheet.max_row
    ncols = sheet.max_column

    height = 1
    for col in range(1, ncols + 1):
        cell = sheet.cell(row=1, column=col)
        _, _, x2, _ = get_merge_cell_size(sheet, cell.coordinate)
        height = max(height, x2)

    return (
        get_sub_sheet(sheet, 1, 1, height, ncols).active,
        get_sub_sheet(sheet, height + 1, 1, nrows, ncols).active,
    )


def split_schema_column(sheet):
    """给定sheet, 以sheet对格式返回schema和剩下的数据"""
    nrows = sheet.max_row
    ncols = sheet.max_column
    width = 1
    for row in range(1, nrows + 1):
        cell = sheet.cell(row=row, column=1)
        _, _, _, y2 = get_merge_cell_size(sheet, cell.coordinate)
        width = max(width, y2)

    return (
        get_sub_sheet(sheet, 1, 1, nrows, width).active,
        get_sub_sheet(sheet, 1, width + 1, nrows, ncols).active,
    )


def main():
    pass


if __name__ == "__main__":
    main()
